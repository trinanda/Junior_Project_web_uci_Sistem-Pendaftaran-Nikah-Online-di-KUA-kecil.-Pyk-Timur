from openpyxl import Workbook

wb = Workbook()

# grab the active worksheet
ws = wb.active

# data can be assigned directly to cells
ws['A1'] = 43

# rows can also be appended
ws.append([1])
ws.append([2])
ws.append([5])

# python types will automatically be converted
# import datetime
# ws['A3'] = datetime.datetime.now()

# save the file
wb.save('testing.xlsx')